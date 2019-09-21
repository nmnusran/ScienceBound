# Naufer Nusran.  2019/09/21
# Following codes were put-together for a demo at the Ames Lab's Science Bound
# AM2302 (DHT22) temperature/humidity sensor is connected to a Raspberry Pi
# All meauserements are appended to a spreadsheet file.


import Adafruit_DHT, time, datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook


numofptns = 10  # number of measurements
waittime =  2   # waittime between measurements
rfilename =  "data.xlsx" # spreadsheet name

##############################################################################
pin = 18    # AM2302 sig connected to GPIO18 
sensor = Adafruit_DHT.DHT22


try:
    wb = load_workbook(rfilename)
    ws = wb.worksheets[0] 
except:
    wb = Workbook()
    ws = wb.active
    header=['Time', 'Temperature (F)', 'Humidity (%)']
    ws.append(header)

starttime=dt.datetime.now().strftime('%H:%M:%S')
print 'start time = ' + starttime


for t in range(numofptns):
    try:
        hum, temp= Adafruit_DHT.read_retry(sensor, pin)   
        temp_f = round(temp * 1.8 + 32.0, 2) 
        humidity = round(hum,1)
        timenow=dt.datetime.now().strftime('%H:%M:%S')
        print str(t)+ "  :  " +str(temp_f) + " F ,  " + str(humidity) + " %"
        time.sleep(waittime)    # Wait before sampling data again
        results = [timenow, temp_f, humidity]
        ws.append(results)
        wb.save(rfilename)

    except RuntimeError as error:
        print error.args[0]

endtime=dt.datetime.now().strftime('%H:%M:%S')

print 'end time = ' + endtime
